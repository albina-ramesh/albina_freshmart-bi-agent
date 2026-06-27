"""
FreshMart – Schritt 1: Daten erstellen
Erstellt 10'000 Verkaufsdatensätze und speichert sie als Excel-Datei.
"""

import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)
np.random.seed(42)

# ── Stammdaten ────────────────────────────────────────────
FILIALEN = {
    "ZH-City":       {"kanton": "ZH", "groesse": "gross",  "typ": "urban"},
    "ZH-Oerlikon":   {"kanton": "ZH", "groesse": "mittel", "typ": "gemischt"},
    "BS-Innenstadt": {"kanton": "BS", "groesse": "mittel", "typ": "urban"},
    "BE-Bern":       {"kanton": "BE", "groesse": "gross",  "typ": "gemischt"},
    "LU-Luzern":     {"kanton": "LU", "groesse": "klein",  "typ": "touristisch"},
    "SG-StGallen":   {"kanton": "SG", "groesse": "mittel", "typ": "gemischt"},
    "GE-Genf":       {"kanton": "GE", "groesse": "gross",  "typ": "urban"},
    "AG-Aarau":      {"kanton": "AG", "groesse": "klein",  "typ": "suburban"},
}

PRODUKTE = {
    "Vollmilch 1L":       ("Milchprodukte",  1.85, 0.18),
    "Bio-Yogurt":         ("Milchprodukte",  0.95, 0.25),
    "Butter 250g":        ("Milchprodukte",  2.40, 0.20),
    "Emmentaler 200g":    ("Milchprodukte",  3.90, 0.28),
    "Weissbrot":          ("Backwaren",      2.10, 0.35),
    "Vollkornbrot":       ("Backwaren",      3.20, 0.38),
    "Gipfeli 4er":        ("Backwaren",      2.80, 0.42),
    "Rindfleisch 400g":   ("Fleisch",        9.90, 0.22),
    "Pouletbrust 500g":   ("Fleisch",        7.50, 0.20),
    "Lachs 300g":         ("Fisch",         11.90, 0.25),
    "Thon Dose":          ("Fisch",          2.30, 0.30),
    "Äpfel 1kg":          ("Früchte",        3.20, 0.32),
    "Bananen 1kg":        ("Früchte",        1.90, 0.28),
    "Erdbeeren 500g":     ("Früchte",        4.50, 0.35),
    "Tomaten 500g":       ("Gemüse",         2.40, 0.30),
    "Karotten 1kg":       ("Gemüse",         1.60, 0.35),
    "Kopfsalat":          ("Gemüse",         1.80, 0.38),
    "Pasta 500g":         ("Trockenwaren",   1.40, 0.22),
    "Reis 1kg":           ("Trockenwaren",   2.20, 0.20),
    "Olivenöl 500ml":     ("Trockenwaren",   8.90, 0.28),
    "Orangensaft 1L":     ("Getränke",       2.90, 0.25),
    "Mineralwasser 1.5L": ("Getränke",       0.85, 0.15),
    "Cola 1.5L":          ("Getränke",       2.50, 0.20),
    "Rotwein Flasche":    ("Alkohol",       12.90, 0.35),
    "Bier 6er-Pack":      ("Alkohol",        9.80, 0.30),
    "Schokolade 100g":    ("Süsswaren",      2.20, 0.40),
    "Chips 150g":         ("Süsswaren",      3.10, 0.38),
    "Kaffee 250g":        ("Heissgetränke",  6.90, 0.32),
    "Teebeutel 20er":     ("Heissgetränke",  3.50, 0.35),
    "Waschmittel 1kg":    ("Haushalt",       7.90, 0.18),
}

ZAHLUNGSMETHODEN = ["Karte", "Bargeld", "TWINT", "Rechnung"]
WOCHENTAGE = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]
FILIAL_GEWICHT = [0.20, 0.15, 0.10, 0.18, 0.08, 0.10, 0.14, 0.05]

# ── Generierung ───────────────────────────────────────────
print("Generiere 10'000 Verkaufsdatensätze...")

START = datetime(2024, 1, 1)
datumsliste = [START + timedelta(days=i) for i in range(366)]
filial_namen = list(FILIALEN.keys())
produkt_namen = list(PRODUKTE.keys())

records = []
for i in range(10000):
    datum   = random.choice(datumsliste)
    filiale = random.choices(filial_namen, weights=FILIAL_GEWICHT)[0]
    produkt = random.choice(produkt_namen)
    kat, preis_basis, marge = PRODUKTE[produkt]

    # Wochentag-Faktor
    wf = [0.85,0.88,0.90,0.92,1.10,1.35,0.70][datum.weekday()]
    # Saison-Faktor
    sf = 1.25 if datum.month in [11,12] else (1.10 if datum.month in [6,7,8] else 1.0)

    menge  = max(1, min(20, int(random.randint(1,8) * sf * wf * random.uniform(0.8,1.2))))
    preis  = round(preis_basis * random.uniform(0.95, 1.05), 2)
    umsatz = round(preis * menge, 2)
    gewinn = round(umsatz * marge, 2)
    zahlung = random.choices(ZAHLUNGSMETHODEN, weights=[0.55,0.20,0.22,0.03])[0]
    stunde  = random.choices(range(7,21), weights=[2,4,6,8,10,12,10,10,10,8,8,6,5,3])[0]

    records.append({
        "Transaktion_ID":   f"TXN-{i+1:05d}",
        "Datum":            datum.strftime("%Y-%m-%d"),
        "Wochentag":        WOCHENTAGE[datum.weekday()],
        "Monat":            datum.strftime("%B"),
        "Quartal":          f"Q{(datum.month-1)//3+1}",
        "Uhrzeit":          f"{stunde:02d}:{random.randint(0,59):02d}",
        "Filiale":          filiale,
        "Kanton":           FILIALEN[filiale]["kanton"],
        "Filialgroesse":    FILIALEN[filiale]["groesse"],
        "Filialtyp":        FILIALEN[filiale]["typ"],
        "Produkt":          produkt,
        "Kategorie":        kat,
        "Menge":            menge,
        "Preis_CHF":        preis,
        "Umsatz_CHF":       umsatz,
        "Gewinn_CHF":       gewinn,
        "Marge_Pct":        round(marge * 100, 1),
        "Zahlungsmethode":  zahlung,
    })

df = pd.DataFrame(records)

# ── Excel erstellen ───────────────────────────────────────
print("Erstelle Excel-Datei mit Formatierung...")

DATEINAME = "freshmart_verkaufsdaten.xlsx"

with pd.ExcelWriter(DATEINAME, engine="openpyxl") as writer:
    # Sheet 1: Rohdaten
    df.to_excel(writer, sheet_name="Verkaufsdaten", index=False)

    # Sheet 2: Zusammenfassung nach Filiale
    filial_summary = df.groupby("Filiale").agg(
        Transaktionen=("Transaktion_ID","count"),
        Umsatz_CHF=("Umsatz_CHF","sum"),
        Gewinn_CHF=("Gewinn_CHF","sum"),
        Avg_Bon_CHF=("Umsatz_CHF","mean")
    ).round(2).reset_index()
    filial_summary["Marge_%"] = ((filial_summary["Gewinn_CHF"] / filial_summary["Umsatz_CHF"]) * 100).round(1)
    filial_summary.to_excel(writer, sheet_name="Filialen", index=False)

    # Sheet 3: Zusammenfassung nach Kategorie
    kat_summary = df.groupby("Kategorie").agg(
        Transaktionen=("Transaktion_ID","count"),
        Umsatz_CHF=("Umsatz_CHF","sum"),
        Gewinn_CHF=("Gewinn_CHF","sum"),
        Menge_Total=("Menge","sum")
    ).round(2).reset_index()
    kat_summary["Marge_%"] = ((kat_summary["Gewinn_CHF"] / kat_summary["Umsatz_CHF"]) * 100).round(1)
    kat_summary.to_excel(writer, sheet_name="Kategorien", index=False)

    # Sheet 4: Monatstrend
    monat_order = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    mon = df.groupby("Monat").agg(
        Transaktionen=("Transaktion_ID","count"),
        Umsatz_CHF=("Umsatz_CHF","sum"),
        Gewinn_CHF=("Gewinn_CHF","sum")
    ).reindex(monat_order).round(2).reset_index()
    mon.to_excel(writer, sheet_name="Monatstrend", index=False)

# ── Formatierung ──────────────────────────────────────────
wb = load_workbook(DATEINAME)

# Farben
GRUEN_DUNKEL = "1A5276"   # Header
GRUEN_HELL   = "D5F5E3"   # Zebra gerade Zeile
WEISS        = "FFFFFF"
GELB         = "F9E79F"   # Titel

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def formatiere_sheet(ws, titel):
    # Titel-Zeile oberhalb einfügen
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")
    titel_cell = ws["A1"]
    titel_cell.value = titel
    titel_cell.font = Font(bold=True, size=13, color=WEISS)
    titel_cell.fill = PatternFill("solid", start_color=GRUEN_DUNKEL)
    titel_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header (Zeile 2)
    for cell in ws[2]:
        cell.font = Font(bold=True, color=WEISS, size=10)
        cell.fill = PatternFill("solid", start_color="2E86C1")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Datenzeilen formatieren
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=1):
        bg = GRUEN_HELL if row_idx % 2 == 0 else WEISS
        for cell in row:
            cell.fill = PatternFill("solid", start_color=bg)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            # CHF-Spalten
            if cell.column_letter in ["P","Q","R","C","D","E"] and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # Spaltenbreite automatisch
    for col in ws.columns:
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
                break
        if not col_letter:
            continue
        max_len = max((len(str(c.value)) if c.value and hasattr(c, 'value') else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

# Formatiere alle Sheets
formatiere_sheet(wb["Verkaufsdaten"], "FreshMart Schweiz – Verkaufsdaten 2024 (10'000 Transaktionen)")
formatiere_sheet(wb["Filialen"],      "FreshMart – Filialübersicht 2024")
formatiere_sheet(wb["Kategorien"],    "FreshMart – Kategorienanalyse 2024")
formatiere_sheet(wb["Monatstrend"],   "FreshMart – Monatstrend 2024")

wb.save(DATEINAME)

print(f"\n✅ Excel-Datei '{DATEINAME}' erfolgreich erstellt!")
print(f"   Sheets: Verkaufsdaten | Filialen | Kategorien | Monatstrend")
print(f"   Datensätze: {len(df):,}")
print(f"   Gesamtumsatz: CHF {df['Umsatz_CHF'].sum():,.2f}")
print(f"\n→ Jetzt Schritt 2 ausführen: python schritt2_datenbank.py")
